"""
Excel Export Service
Handles exporting data to formatted Excel workbooks
"""

from __future__ import annotations
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional, List
import uuid

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    # Mock classes for when openpyxl is not available to avoid NameErrors at class definition time
    class PatternFill:
        def __init__(self, **kwargs): pass
    class Font:
        def __init__(self, **kwargs): pass
    class Border:
        def __init__(self, **kwargs): pass
    class Side:
        def __init__(self, **kwargs): pass
    class Alignment:
        def __init__(self, **kwargs): pass


class ExcelExportService:
    """Service for exporting data to Excel format."""

    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @staticmethod
    def _check_openpyxl_available():
        """Check if openpyxl is installed."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel exports. Install with: pip install openpyxl"
            )

    @staticmethod
    def _style_header_row(ws, row_num: int, num_columns: int):
        """Apply header styling to a row."""
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = ExcelExportService.HEADER_FILL
            cell.font = ExcelExportService.HEADER_FONT
            cell.border = ExcelExportService.BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _style_data_row(ws, row_num: int, num_columns: int):
        """Apply data row styling."""
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = ExcelExportService.BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    @staticmethod
    def _set_column_widths(ws, columns: List[tuple]):
        """
        Set column widths and formatting.
        columns: List of (column_letter, width, format_type) tuples
        """
        for col_letter, width, format_type in columns:
            ws.column_dimensions[col_letter].width = width
            
            # Apply formatting to entire column
            if format_type == "currency":
                for row in ws.iter_rows(min_col=ord(col_letter) - ord('A') + 1,
                                       max_col=ord(col_letter) - ord('A') + 1):
                    for cell in row:
                        if cell.row > 1:  # Skip header
                            cell.number_format = '"$"#,##0.00'
            elif format_type == "percent":
                for row in ws.iter_rows(min_col=ord(col_letter) - ord('A') + 1,
                                       max_col=ord(col_letter) - ord('A') + 1):
                    for cell in row:
                        if cell.row > 1:
                            cell.number_format = '0.00%'
            elif format_type == "number":
                for row in ws.iter_rows(min_col=ord(col_letter) - ord('A') + 1,
                                       max_col=ord(col_letter) - ord('A') + 1):
                    for cell in row:
                        if cell.row > 1:
                            cell.number_format = '#,##0'

    @staticmethod
    async def export_sales_by_month(
        db: AsyncSession,
        organization_id: uuid.UUID,
        year: int,
        month: Optional[int] = None,
    ) -> BytesIO:
        """
        Export sales data to Excel.
        
        If month is provided: Creates single sheet with that month's sales
        If month is None: Creates 12-sheet workbook with one sheet per month
        """
        ExcelExportService._check_openpyxl_available()
        
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        if month:
            # Single month export
            ws = wb.create_sheet(f"Sales {month:02d}")
            await ExcelExportService._populate_sales_sheet(
                db, ws, organization_id, year, month
            )
        else:
            # Full year export - 12 sheets
            for m in range(1, 13):
                ws = wb.create_sheet(f"{datetime(year, m, 1).strftime('%B')}")
                await ExcelExportService._populate_sales_sheet(
                    db, ws, organization_id, year, m
                )
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def _populate_sales_sheet(
        db: AsyncSession,
        ws,
        organization_id: uuid.UUID,
        year: int,
        month: int,
    ):
        """Populate a worksheet with sales data for a specific month."""
        # Note: This is a placeholder implementation
        # In a full implementation, query database for sales data matching the period
        
        headers = ["Date", "Branch", "Cashier", "Customer", "Items", "Total", "Discount", "Tax"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ExcelExportService._style_header_row(ws, 1, len(headers))
        ExcelExportService._set_column_widths(
            ws,
            [("A", 12, "text"), ("B", 15, "text"), ("C", 12, "text"), 
             ("D", 15, "text"), ("E", 8, "number"), ("F", 12, "currency"),
             ("G", 12, "currency"), ("H", 12, "currency")]
        )
        
        # Add sample data structure (actual data would come from database query)
        # This ensures the sheet is properly formatted even if no data

    @staticmethod
    async def export_inventory_by_month(
        db: AsyncSession,
        organization_id: uuid.UUID,
        year: int,
        month: int,
    ) -> BytesIO:
        """Export inventory snapshot to Excel."""
        ExcelExportService._check_openpyxl_available()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Inventory {month:02d}"
        
        headers = ["Drug Name", "Branch", "Quantity on Hand", "Unit Price", "Total Value", "Batch #", "Expiry Date"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ExcelExportService._style_header_row(ws, 1, len(headers))
        ExcelExportService._set_column_widths(
            ws,
            [("A", 20, "text"), ("B", 15, "text"), ("C", 12, "number"),
             ("D", 12, "currency"), ("E", 12, "currency"), ("F", 12, "text"), ("G", 12, "text")]
        )
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def export_staff_data(
        db: AsyncSession,
        organization_id: uuid.UUID,
    ) -> BytesIO:
        """Export staff/user directory to Excel."""
        ExcelExportService._check_openpyxl_available()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Staff Directory"
        
        headers = ["Name", "Email", "Phone", "Role", "Branch", "Status", "Created Date"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ExcelExportService._style_header_row(ws, 1, len(headers))
        ExcelExportService._set_column_widths(
            ws,
            [("A", 18, "text"), ("B", 20, "text"), ("C", 15, "text"),
             ("D", 12, "text"), ("E", 15, "text"), ("F", 12, "text"), ("G", 12, "text")]
        )
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
